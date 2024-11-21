# -*- coding: utf-8 -*-
"""
Created on Wed Sep 30 12:26:11 2020

@author: VanTongeren_F

metropy - Module to convert OECD METRO results GDX to Python using pandas dataframe
"""
# standard library imports
import os
import sys
import shutil

# third party imports
import gdxpds          # module to read gdx into pandas, see https://github.com/NREL/gdx-pandas
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


#%% classes
class metro_data(object):
    '''
    class to wrap METRO data file containing results
    '''
    def __init__(self, dataID):
        self.dataID = dataID
        self._data = {}
        self._sets = {}
        self.path = " "
        self.setpath = " "
        self.legend = None # use to describe data
        self._fullname = {}

    @property
    def fullname(self):
        self._fullname[self.dataID] = self.legend
        return self._fullname

    @property
    def data(self):
        ''' Contains a METRO results file
        Loads file on <path> from disk if not already loaded
        '''
        if self._data:
            return self._data #data already read, return directly

        try:
            print(f'Loading GDX file {self.path}\n')
            self._data = gdxpds.to_dataframes(self.path)
            self._data['results'].columns=['dim1','dim2','variable','dim4','dim5','value']
            #NOTE: the META and META_p information is not read in correctly.
            #This is a type conversion issue in gdxpds. So delete it from dict until solution found
            if 'META'in self._data.keys():  del self._data['META']
            if 'META_p'in self._data.keys(): del self._data['META_p']

        except gdxpds.Error as msg:
            print(f'{msg} \n'
                  f'Please use metro_obj.path = <full path to gdx file containing METRO results>\n')

            sys.exit()

        return self._data

    @property
    def sets(self):
        if self._sets:
            return self._sets
        sets_file = ''
        try:
            sets_file = pd.ExcelFile(os.path.join(self.setpath) )
        except FileNotFoundError as msg:
            print(f'{msg} \nCannnot find sets file. '
                  f'Please use metro_obj.setpath=<full path to xls setsfile>'
                  f'\nUsing short set names as default instead.\n')

            self._sets['rregions'] = dict(zip(self.dimensions['rregions'],
                                       self.dimensions['rregions']))
            self._sets['wregions'] = self._sets['rregions']
            self._sets['factors'] = dict(zip(self.dimensions['factors'],
                                       self.dimensions['factors']))
            self._sets['commodities'] = dict(zip(self.dimensions['commodities'],
                                          self.dimensions['commodities']))
            self._sets['activities'] = self._sets['commodities']


        if sets_file:
            print(f'Loading XLSX sets file {self.setpath}\n')

            df_sets=pd.read_excel(sets_file, 'sets-maps', header=None,\
                                 skiprows =[0,1,2,3], usecols=[0,1,6,7,14,15])

            df_sets.dropna(how='all', inplace=True)
            df_sets.columns=['regions', 'regions_descr', \
                            'factors', 'factors_descr', \
                            'sectors', 'sectors_descr']

            regions=df_sets[['regions','regions_descr']].sort_values(by='regions')
            rregions= sorted(self.dimensions['rregions'][:])
            self._sets['rregions'] = dict(zip(rregions, regions['regions_descr']))
            self._sets['wregions'] = self._sets['rregions']

            factors = df_sets[['factors', 'factors_descr']].sort_values(by='factors')
            dfactors = sorted(self.dimensions['factors'][:])
            self._sets['factors'] = dict(zip(dfactors, factors['factors_descr']))

            sectors = df_sets[['sectors', 'sectors_descr']].sort_values(by='sectors')
            commodities= sorted(self.dimensions['commodities'][:])
            self._sets['commodities'] = dict(zip(commodities, sectors['sectors_descr']))
            self._sets['activities'] = self._sets['commodities']

        return self._sets


    @property
    def dimensions(self):
        '''returns a dict of all the dimensions in the METRO results file
        By default, the current METRO results parameter is a 5-dimensional array, where
        the contents of each dimension can vary.

        rregions    : (origin) regions
        wregions    : (destination) regions
        factors     : production factors
        variables   : variables
        commodities : commodities
        activities  : actvities
        usecat      : use categories
        otherdims   : other dimensions

        Example to get the list of variables
                    metro_data_object.dimensions[ 'variables']
        '''
        res= self.data['results']

        rregions = list(pd.unique(res['dim1']))
        if 'empty'in rregions: rregions.remove('empty')
        wregions = [s for s in pd.unique(res['dim2']) if s[0]=='w']
        factors = [s for s in pd.unique(res['dim2']) if s[0]=='f']
        variables = list(pd.unique(res['variable']))
        commodities = [s for s in pd.unique(res['dim4']) if s[0]=='c']
        activities = [s for s in pd.unique(res['dim4']) if s[0]=='a']
        usecat = [s for s in pd.unique(res[ 'dim5']) if s[0]=='u']
        otherdims = [s for s in pd.unique(res['dim5']) if s not in usecat]
        otherdims.remove('empty')

        return {'variables':variables, \
                'rregions':rregions, 'wregions': wregions, \
                'factors':factors, \
                'commodities': commodities, 'activities':activities, \
                'usecat': usecat, \
                'otherdims':otherdims}

    def get_set(self, s):
        ''' retuns a dict of the set s as {member_name: member_description}
        to convert to a list: list(metro_obj.get_sets(setname))'''
        if not self._sets: # not loaded yet
            self.sets      # call sets property will load
        try:
            return self._sets[s]
        except KeyError:
            print(f'* ERROR * There is no set "{s}" in the sets file \n'
                  f'  Try one of the following set names: {self._sets.keys()}')


    @property
    def _variables(self):
        ''' returns a dict of dataframes by variable '''
        return dict(list(self.data['results'].groupby(['variable'])))

    def get_variable(self, v):
        '''returns values of one variable '''
        dims=['dim1', 'dim2','dim4','dim5']
        try:
            tmp = self._variables[v]
            for c in tmp.columns: # remove columns that only have 'empty' as value
                if (tmp[c] == 'empty').all():
                   dims.remove(c)

            return tmp.groupby(dims).sum().unstack()

        except KeyError:
            print(f'* ERROR * There is no variable "{v}" in the results file')

    def __str__(self):
        s = str(self.fullname).strip("{" "}")
        s += ", " + self.path
        return s

#end class metro_data()

class result_stack(object):
    ''' Class containing stack of metro_data objects
    '''
    def __init__(self):
        self.stack = []

    def add_result(self, metro_obj):
        ''' Adds metro_data object to stack'''
        self.stack.append(metro_obj)

    def get_results(self):
        '''Returns a list of metro_data objects in stack'''
        return self.stack[:] # return copy of stack

    def get_dataIDs(self):
        '''Returns list of dataIDs of all resukts in stack '''
        return [s.dataID for s in self.stack]

    def get_fullnames(self):
        '''Returns list of fullnames of all results in stack '''
        return [s.fullname for s in self.stack]


    def get_var(self, v):
        ''' Gets one variable from all metro_data objects in stack
        To find out which metro results are in stack use:
            <.names> attribute'''
        df_dict = {}
        for k in self.stack:
            df_dict[k.dataID]= k.get_variable(v)
        return df_dict

    def __str__(self):
        pass # nothing to print by default

#end class result_stack

#%% Helper functions for dataframes and series
def add_ctotal(df, name='TOTAL'):
    '''Adds column totals to dataframe or series df
    NOTE: the df.append method alters the index of the df '''
    c_total=[]
    if is_series(df):
        c_total = pd.Series(df.sum(), index=[name])
        return df.append(c_total)
    else:
        c_total = pd.Series(df.sum(), index=df.columns, name=name)

    return df.append(c_total)


def add_rtotal(df, name='TOTAL'):
    '''Adds row totals to dataframe df
    If df is a Series it adds a column total'''
    if is_series(df):
        return add_ctotal(df, name)
    else:
        df[name] = df.sum(axis=1)
        return df

def add_mapper(df, col, mapper):
    ''' Adds a column to dataframe df to map col into a new list
    df: dataframe; col: a valid column name or index level name; mapper: dict
    If the passed df is result of groupby operations, col may be found in the multi-index
    Returns: df with a column added named 'col_map'
    Example: add_mapper(my_df, 'col1', map_dict) '''

    idx=df.index # remember old multi-index
    map_nm = str(col)+'_map'
    if  col in df.columns:
        df[map_nm]= df[col].map(mapper)

    elif col in df.index.names:
        #level=df.index.names.index(col)
        df.reset_index(level=col, inplace=True, drop=False) #put 'col'from index into df column
        df[map_nm]= df[col].map(mapper) # apply mapping
        df.set_index(idx, inplace=True)
        df.drop(columns=col, inplace=True)
    else:
        raise KeyError(f'* ERROR * name not found in columns or index')

    return df
# end add_mapper()

def pct_diff(df1, df2):
    ''' Returns percent difference between df1 and df2 for all valid columns
    Accepts Pandas series and Pandas dataframes
    Ignores nun-numeric columns, i.e. not 'float64' or 'int64'
    Returns a dataframe: (df2/df1 -1)*100
    '''
    # convert series to df if necessary:
    if is_series(df1):
        df1 = pd.DataFrame(df1)
    if is_series(df2):
        df2 = pd.DataFrame(df2)

    assert len(df1.columns) == len(df2.columns)     # same no of columns
    assert 'False' not in list(df1.columns == df2.columns) # all column names the same

    # check if columns have non-numerical values and drop if necessary:
    to_drop=[]
    for c in range(len(df1.columns)):
        if df1.iloc[:,c].dtypes not in ['float64', 'int64']:   #not a valid data type
            to_drop.append(c)

    d1=df1.drop(df1.columns[(to_drop)], axis = 1, inplace=False)
    d2=df2.drop(df2.columns[(to_drop)], axis = 1, inplace=False)

    return (d2 / d1 -1.0)*100
#end def pct_change

def pct_diff_list(base_df, df_list, headers=[" "], basecol= False):
    ''' Returns percent difference between base_df and the
    data frames in the list df_list, for all columns.
    Accepts Pandas series and Pandas dataframes
    Returns a dataframe with a column multi-index:
        (elements in df_list,(original index of base_df))
        i.e: it puts the percent differences for each frame in df_list side-by-side
    base_df: dataframe taken as base
    df_list: a list of datafranes - must be of same layout as base_df
    headers:  optional list of headers to be used in resulting df; must have length
             len(df_list) +1
    basecol: boolean, if True and base_df is a series it inserts base_df at the
             beginning of dataframe, ignored if base_df has more than one column
    '''
 #TODO use dataID from metro_data class?
    # create values for new columns index for output df
    if len(headers)-1 != len(df_list):
        #list of headers is empty or incomplete, so just use sequential numbers
        cols= list(range(len(df_list)))
    else:
        cols=headers[1:]

    #create new column index for output df:
    if is_series(base_df):
        #take care of series - they have no columns defined
        idx= cols
    else:
        idx_b= list(base_df.columns)
        idx=pd.MultiIndex.from_product([cols,idx_b]) # kronecker product of cols and idx_b

    res = pd.DataFrame(columns=idx, index= base_df.index)

    #actual calculations here:
    for k,c in enumerate(cols):
        res[c] = pct_diff(base_df, df_list[k])

    if basecol and is_series(base_df):
        # insert base values at left of output df
        res.insert(0,'Base values',base_df)

    return res
#end def pct_change_list

def is_series(df):
    '''returns True if df is type pandasSeries '''
    return str(type(df)) == "<class 'pandas.core.series.Series'>"
#end def is_series()

#standard macro table
def macro_table(M_data, longnames=False):
    ''' Standard METRO macro results table

    PARAMETERS
    -----
    M_data: a metro.metro_data() object
    longname: boolean, if True variable and region longnames are used

    RETURNS
    ---
    A standard table with macro variables in the rows and regions in the columns
    '''

    variables = {'rGDPEXP' : 'real GDP',
                 'rQABSORP': 'Domestic absorption',
                 'rQCDTOT' : 'Consumption demand',
                 'rQGDTOT' : 'Government demand',
                 'rQINVTOT': 'Investment demand' ,
                 'rEXPORT' : 'Exports',
                 'rIMPORT' : 'Imports' }

    df_tmp=pd.DataFrame(columns=['rregion','value', 'variable', 'var_descript'])

    for v in variables.keys():
        var= M_data.get_variable(v)
        df=pd.DataFrame(var, columns=['value'])
        df['variable'] = v
        df.reset_index(inplace=True)
        df.drop(columns='level_0', inplace=True)
        df.rename(columns={'dim1': 'rregion'}, inplace=True)
        df_tmp=df_tmp.append(df, sort=False)

    if longnames:
        df_tmp['reg_longname']= df_tmp['rregion'].map(M_data.get_set('rregions'))
        df_tmp['var_descript'] = df_tmp['variable'].map(variables)
        df_out=df_tmp.groupby(by=['var_descript','variable', 'reg_longname'], \
                          sort=False).sum().unstack('reg_longname')
    else:
        df_out=df_tmp.groupby(by=['variable', 'rregion'], \
                          sort=False).sum().unstack('rregion')

    df_out = add_rtotal(df_out)

    return df_out
#end macro_table

def out_tables():
    table_dict={}
    table_dict['ReadMe'] = ['put ReadMe info here as a list:',
                            '"[line 1, line 2, ...]" ']
    return table_dict
#end out_table

def add_to_out_tables(table_name, df, sheet_name = '', info=''):
    ''' Adds df to the dict table_name, giving it optional sheet_name and info string
    if no sheet_name is given a default name is created from the shape of df
    Note: if there are multiple df with the same shape, only the last one will be
    preserved
    '''
    if sheet_name == '': sheet_name= 'Table '+str(df.shape)
    table_name[sheet_name] = df, info
#end add_to_out_tables


def make_wksheet(df, wb, sheet_name = 'RESULT', info=' '):
    ''' Makes Excel worksheet from df in workbook wb '''

    def make_header_rows(df):
        ''' make table header rows
            need to do some ugly stuff because openpyxl can't convert complex column multiindex.
            NOTE: in some cases this can still fail to work, especially if c_totals
            are added and the df index gets mangled
        '''
        idx_size = df.columns.size
        idx_levels = df.columns.nlevels
        rows = {}
        for l in range(idx_levels):
            r = []
            for i in range(idx_size):
                element = df.columns.get_level_values(l)[i]
                size_element = len(df.columns.get_level_values(l)[i])
                if type(element) != tuple:
                    r.append(element)
                else: #tuple
                    for k in range(size_element):
                        element = df.columns.get_level_values(l)[i][k]
                        if element not in ('value', '') : r.append(element)

            rows[l] = r

        return rows
    #end make_header_rows()

    def get_max_tupsize(idx):
        n=1
        for r, v in enumerate(idx):
            if type(v) == tuple:
                    if len(v) > n: n = len(v)
        return n

    def strip_index(df):
        ''' strip multi index and put it in columns.
        This is necessary, as xls does not understand tuples as contents
        of spreadsheet cells'''

        df_work = df.copy()
        nl = get_max_tupsize(df_work.index)

        if nl == 1: # single level index
            df_work.reset_index(inplace=True)
        else:       # multilevel index
            for c in range(nl):
                temp=[]
                for r, v in enumerate(df_work.index):
                    if type(v) == tuple:
                        temp.append(v[c])
                    else:
                       temp.append(v)
                df_work['index_'+str(c)] = temp

            swapped=[]
            new_cols=(df_work.columns)
            for c in range(-nl, 0):
                swapped.append(new_cols[c])

            for c in range(len(new_cols) - nl):
                swapped.append(new_cols[c])

            if df_work.columns.nlevels > 1:
                  new_colIDX = pd.MultiIndex.from_tuples(swapped)
            else:
                  new_colIDX = swapped

            df_work = df_work.reindex(columns=new_colIDX)
            df_work.reset_index(drop=True)
        #end else
        return df_work
    # end strip

    ws = wb.create_sheet(title = sheet_name)
    ws['A1'] = info

    #df.reset_index(inplace=True) # reset multi index for openpyxl to understand
    #df.set_index(df.columns[0][0], inplace=True, drop=True)

    df = strip_index(df)
    
    head = make_header_rows(df)
    
    for i in range(len(head)): ws.append(head[i])
    
    for r in dataframe_to_rows(df, index=False, header=False):
        
        try:
            ws.append(r)
        except ValueError:
            print(f'Cannot write to excel row: {r}')
            pass

    # some cell formatting
    for cell in ws['A'] + ws[2] +ws[3]:
        cell.font = Font(bold=True)
    ws['A1'].font= Font(color="FF0000", bold=True) #red bold

    return ws
#end make_wksheet

def write_to_excel(tables, filename):
    ''' Writes a dict of tables to excel file
        Also writes a ReadMe sheet if present in the table dict'''

    wb = Workbook()
    ws_list=list()

    # create tables output directory if necessary
    if not os.path.isdir(os.path.dirname(filename)):
         os.makedirs(os.path.dirname(filename))

    #backup existing file
    fn_bak=filename[:-5] + '~' + filename[-5:]
    if os.path.isfile(filename):
        try:                                 # make backup copy of log file
            shutil.copyfile(filename, fn_bak)
        except shutil.SameFileError:
            pass

    for i, s in enumerate(tables.keys()):
        if str(s).upper() == 'README': # add readme sheet
            ws_list.append(wb.create_sheet(title = 'ReadMe', index=0))
            for i, v in enumerate(list(tables[s])):
                _ = ws_list[-1].cell(column= 1, row = i+1, value = v)
        else: #add table sheet
            ws_list.append(make_wksheet(tables[s][0], wb, sheet_name = s, info = tables[s][1]))
    try:
        wb.save(filename = filename)
    except PermissionError as msg:
        print(f'The xls file to write your tables is probably open. '
              f'Please close it first.\n{msg}')

#end write-to_excel

def main():
    pass
if __name__ == "__main__":
    main()




